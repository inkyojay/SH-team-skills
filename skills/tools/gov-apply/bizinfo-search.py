#!/usr/bin/env python3
"""기업마당 API 기반 지원사업 검색 및 적합도 분석 도구 v2

사용법:
  python3 bizinfo-search.py <기업정보.json> [옵션]

옵션:
  --all          마감 전 전체 사업 출력 (기본: 적합도 30점 이상만)
  --min-score N  최소 적합도 점수 (기본: 0 with --all, 30 without)
  --output PATH  결과 저장 경로
  --xlsx PATH    엑셀로 저장
"""

import json, sys, os, urllib.request, urllib.parse, argparse
from datetime import datetime, date
from html import unescape
import re

API_KEY = "4389Yp"
API_URL = "https://www.bizinfo.go.kr/uss/rss/bizinfoApi.do"

def fetch_programs(num=500):
    params = urllib.parse.urlencode({
        "crtfcKey": API_KEY,
        "dataType": "json",
        "numOfRows": num,
    })
    url = f"{API_URL}?{params}"
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))

def clean_html(text):
    if not text:
        return ""
    text = unescape(text)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def parse_date_range(date_str):
    try:
        parts = date_str.split("~")
        start = parts[0].strip()[:10]
        end = parts[1].strip()[:10] if len(parts) > 1 else start
        return (
            datetime.strptime(start, "%Y-%m-%d").date(),
            datetime.strptime(end, "%Y-%m-%d").date()
        )
    except:
        return None, None

def score_program(item, company_info):
    score = 0
    reasons = []
    name = item.get("pblancNm", "")
    summary = clean_html(item.get("bsnsSumryCn", ""))
    target = item.get("trgetNm", "")
    hashtags = item.get("hashtags", "")
    field_l = item.get("pldirSportRealmLclasCodeNm", "")
    field_m = item.get("pldirSportRealmMlsfcCodeNm", "")
    combined = f"{name} {summary} {hashtags} {field_l} {field_m}".lower()

    # 대상 기업 유형
    if "중소기업" in target or "소기업" in target or "소상공인" in target:
        score += 20
        reasons.append("중소기업 대상")
    elif "중견기업" in target:
        score += 10
        reasons.append("중견기업 대상")

    # 업종 키워드 (중복 가산)
    industry_hits = []
    industry_keywords = {
        "유아": 30, "아동": 25, "베이비": 30, "키즈": 25, "출산": 25,
        "침구": 30, "섬유": 20, "생활용품": 25, "소비재": 25,
        "제조": 15, "제조업": 15,
    }
    for kw, pts in industry_keywords.items():
        if kw in combined:
            industry_hits.append((kw, pts))
    if industry_hits:
        best = max(industry_hits, key=lambda x: x[1])
        score += best[1]
        reasons.append(f"업종 '{best[0]}'")

    # 수출/해외
    export_hits = []
    export_keywords = {"수출": 25, "해외": 20, "글로벌": 20, "바이어": 20,
                       "전시회": 15, "유통망": 20, "입점": 15, "무역": 15, "통상": 10}
    for kw, pts in export_keywords.items():
        if kw in combined:
            export_hits.append((kw, pts))
    if export_hits:
        best = max(export_hits, key=lambda x: x[1])
        score += best[1]
        reasons.append(f"수출 '{best[0]}'")

    # 마케팅/디자인
    mkt_hits = []
    mkt_keywords = {"마케팅": 15, "브랜드": 15, "디자인": 15,
                    "홍보": 10, "광고": 10, "온라인": 10, "카탈로그": 15, "판로": 10}
    for kw, pts in mkt_keywords.items():
        if kw in combined:
            mkt_hits.append((kw, pts))
    if mkt_hits:
        best = max(mkt_hits, key=lambda x: x[1])
        score += best[1]
        reasons.append(f"마케팅 '{best[0]}'")

    # 자금 지원
    fund_keywords = {"바우처": 10, "지원금": 10, "보조금": 10}
    for kw, pts in fund_keywords.items():
        if kw in combined:
            score += pts
            reasons.append(f"자금 '{kw}'")
            break

    # 지역 (경기도 용인)
    if "경기" in combined or "경기" in hashtags:
        score += 15
        reasons.append("경기도")
    elif "용인" in combined:
        score += 20
        reasons.append("용인시")
    elif "전국" in combined or "전국" in hashtags:
        score += 10
        reasons.append("전국")

    # 제외 (감점)
    exclude_keywords = ["농업", "농촌", "수산", "축산", "건설업", "조선", "방위산업",
                        "원자력", "자동차부품", "게임콘텐츠", "영화", "음악",
                        "의약품", "바이오", "반도체", "이차전지", "로봇"]
    for kw in exclude_keywords:
        if kw in combined:
            score -= 40
            reasons.append(f"비관련 '{kw}'")
            break

    return max(score, 0), reasons

def grade_label(score):
    if score >= 70: return "최상"
    if score >= 50: return "상"
    if score >= 30: return "중"
    if score >= 10: return "하"
    return "-"

def main():
    parser = argparse.ArgumentParser(description="기업마당 지원사업 검색")
    parser.add_argument("company_info", nargs="?", help="기업정보.json 경로")
    parser.add_argument("--all", action="store_true", help="전체 사업 출력")
    parser.add_argument("--min-score", type=int, default=None, help="최소 적합도")
    parser.add_argument("--output", default=None, help="JSON 저장 경로")
    parser.add_argument("--xlsx", default=None, help="엑셀 저장 경로")
    args = parser.parse_args()

    min_score = args.min_score if args.min_score is not None else (0 if args.all else 30)

    company_info = {}
    if args.company_info and os.path.exists(args.company_info):
        with open(args.company_info, "r") as f:
            company_info = json.load(f)

    print("기업마당 API 조회 중...")
    data = fetch_programs(500)
    programs = data.get("jsonArray", [])
    total = programs[0].get("totCnt", 0) if programs else 0
    print(f"총 {total}건 중 {len(programs)}건 조회\n")

    today = date.today()
    results = []

    for item in programs:
        start_date, end_date = parse_date_range(item.get("reqstBeginEndDe", ""))
        if not end_date or end_date < today:
            continue

        score, reasons = score_program(item, company_info)
        d_day = (end_date - today).days

        results.append({
            "name": item["pblancNm"],
            "score": score,
            "grade": grade_label(score),
            "reasons": reasons,
            "deadline": str(end_date),
            "start_date": str(start_date) if start_date else "",
            "d_day": d_day,
            "target": item.get("trgetNm", ""),
            "field_l": item.get("pldirSportRealmLclasCodeNm", ""),
            "field_m": item.get("pldirSportRealmMlsfcCodeNm", ""),
            "url": item.get("pblancUrl", ""),
            "org_main": item.get("jrsdInsttNm", ""),
            "org_exec": item.get("excInsttNm", ""),
            "contact": item.get("refrncNm", ""),
            "apply_method": item.get("reqstMthPapersCn", ""),
            "apply_url": item.get("rceptEngnHmpgUrl", ""),
            "summary": clean_html(item.get("bsnsSumryCn", ""))[:300],
            "hashtags": item.get("hashtags", ""),
            "pblancId": item.get("pblancId", ""),
        })

    # 적합도순 → 마감일순 정렬
    results.sort(key=lambda x: (-x["score"], x["d_day"]))

    # 필터
    filtered = [r for r in results if r["score"] >= min_score]

    # 콘솔 출력
    print(f"{'='*90}")
    print(f"  제이코프(JAYCORP) 지원사업 검색 결과")
    print(f"  기준일: {today} | 마감 전 {len(results)}건 | 적합도 {min_score}점 이상 {len(filtered)}건")
    print(f"{'='*90}\n")

    for i, r in enumerate(filtered, 1):
        print(f"[{i:3d}] {r['grade']:2s} ({r['score']:3d}점) | D-{r['d_day']:<3d} ({r['deadline']})")
        print(f"      {r['name']}")
        print(f"      {r['field_l']}>{r['field_m']} | {r['target']} | {r['org_exec']}")
        if r['reasons']:
            print(f"      매칭: {', '.join(r['reasons'])}")
        print(f"      {r['url']}")
        print()

    # JSON 저장
    output_path = args.output or "/Users/inkyo/Projects/지원사업 지원/적합사업_검색결과.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(filtered, f, ensure_ascii=False, indent=2)
    print(f"JSON 저장: {output_path} ({len(filtered)}건)")

    # 엑셀 저장
    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "지원사업 검색결과"

            headers = ["#", "적합도", "점수", "D-Day", "마감일", "사업명",
                       "분야", "세부분야", "대상", "수행기관", "매칭사유",
                       "링크", "요약"]
            header_fill = PatternFill('solid', fgColor='1F4E79')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

            for j, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=j, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            grade_colors = {
                "최상": PatternFill('solid', fgColor='C6EFCE'),
                "상": PatternFill('solid', fgColor='FFEB9C'),
                "중": PatternFill('solid', fgColor='FFC7CE'),
                "하": PatternFill('solid', fgColor='D9D9D9'),
                "-": PatternFill('solid', fgColor='F2F2F2'),
            }

            for i, r in enumerate(filtered, 1):
                row = i + 1
                ws.cell(row=row, column=1, value=i)
                cell_grade = ws.cell(row=row, column=2, value=r['grade'])
                cell_grade.fill = grade_colors.get(r['grade'], PatternFill())
                ws.cell(row=row, column=3, value=r['score'])
                ws.cell(row=row, column=4, value=f"D-{r['d_day']}")
                ws.cell(row=row, column=5, value=r['deadline'])
                ws.cell(row=row, column=6, value=r['name'])
                ws.cell(row=row, column=7, value=r['field_l'])
                ws.cell(row=row, column=8, value=r['field_m'])
                ws.cell(row=row, column=9, value=r['target'])
                ws.cell(row=row, column=10, value=r['org_exec'])
                ws.cell(row=row, column=11, value=', '.join(r['reasons']))
                ws.cell(row=row, column=12, value=r['url'])
                ws.cell(row=row, column=13, value=r['summary'][:200])

                for col in range(1, 14):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                    ws.cell(row=row, column=col).font = Font(size=9)

            ws.column_dimensions['A'].width = 4
            ws.column_dimensions['B'].width = 6
            ws.column_dimensions['C'].width = 5
            ws.column_dimensions['D'].width = 6
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 45
            ws.column_dimensions['G'].width = 8
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 18
            ws.column_dimensions['K'].width = 30
            ws.column_dimensions['L'].width = 40
            ws.column_dimensions['M'].width = 50

            ws.auto_filter.ref = f"A1:M{len(filtered)+1}"
            ws.freeze_panes = 'F2'

            wb.save(args.xlsx)
            print(f"엑셀 저장: {args.xlsx}")
        except ImportError:
            print("openpyxl 없어서 엑셀 저장 생략")

if __name__ == "__main__":
    main()
