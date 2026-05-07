#!/usr/bin/env python3
"""
xlsx_writer.py — 결과를 xlsx 1개 (전제품 종합 + 카테고리별 시트)로 출력
"""

from __future__ import annotations

from dataclasses import asdict, is_dataclass
from datetime import date
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("[!] openpyxl 필요: pip install openpyxl")


COLUMNS = [
    "슬러그",
    "카테고리",
    "제품 정식명",
    "1줄 USP",
    "소재",
    "사이즈",
    "컬러",
    "네이버 상품명 (1안)",
    "상품명 길이",
    "메인 키워드",
    "메인 검색량(월)",
    "메인 등급",
    "경쟁도",
    "보조1",
    "보조2",
    "보조3",
    "보조4",
    "태그1",
    "태그2",
    "태그3",
    "태그4",
    "태그5",
    "태그6",
    "태그7",
    "태그8",
    "태그9",
    "태그10",
    "시즌 피크",
    "현재 vs 피크 (%)",
    "다중카테고리",
    "경고",
]


HEADER_FILL = PatternFill(start_color="C8A07C", end_color="C8A07C", fill_type="solid")
HEADER_FONT = Font(name="Pretendard", size=11, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Pretendard", size=10)
GRADE_COLORS = {
    "S": "FFE5DD",  # blush
    "A": "F0E6DA",  # accent-light
    "B": "F5F0EB",
    "C": "FAF7F4",
    "D": "FFFFFF",
}


def _row_from_record(rec: dict[str, Any]) -> list[Any]:
    sub_kws = rec.get("sub_keywords", []) + ["", "", "", ""]
    tags = rec.get("tags", []) + [""] * 10

    return [
        rec.get("slug", ""),
        rec.get("category", ""),
        rec.get("title", ""),
        rec.get("usp_short", ""),
        rec.get("material", ""),
        rec.get("size", ""),
        rec.get("color", ""),
        rec.get("product_name", ""),
        len(rec.get("product_name", "")),
        rec.get("main_keyword", ""),
        rec.get("main_search", 0),
        rec.get("main_grade", ""),
        rec.get("main_competition", ""),
        sub_kws[0],
        sub_kws[1],
        sub_kws[2],
        sub_kws[3],
        tags[0],
        tags[1],
        tags[2],
        tags[3],
        tags[4],
        tags[5],
        tags[6],
        tags[7],
        tags[8],
        tags[9],
        rec.get("peak_month", ""),
        rec.get("current_vs_peak", ""),
        rec.get("multi_categories", ""),
        " | ".join(rec.get("warnings", [])),
    ]


def _style_header(ws):
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "C2"  # 슬러그 + 카테고리 freeze


def _autosize(ws):
    widths = {
        "슬러그": 22,
        "카테고리": 14,
        "제품 정식명": 26,
        "1줄 USP": 28,
        "소재": 22,
        "사이즈": 18,
        "컬러": 14,
        "네이버 상품명 (1안)": 50,
        "상품명 길이": 8,
        "메인 키워드": 18,
        "메인 검색량(월)": 12,
        "메인 등급": 8,
        "경쟁도": 8,
        "다중카테고리": 24,
        "경고": 30,
    }
    for col_idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col, 14)


def _grade_row_fill(ws, row_idx: int, grade: str):
    color = GRADE_COLORS.get(grade, "FFFFFF")
    if color == "FFFFFF":
        return
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def write_xlsx(records: list[dict[str, Any]], output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # 시트 1: 전제품 종합
    ws_all = wb.create_sheet("전제품 종합")
    _style_header(ws_all)
    for i, rec in enumerate(records, 2):
        row = _row_from_record(rec)
        for col_idx, val in enumerate(row, 1):
            cell = ws_all.cell(row=i, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        _grade_row_fill(ws_all, i, rec.get("main_grade", ""))
    _autosize(ws_all)

    # 시트 2~N: 카테고리별
    by_cat: dict[str, list[dict[str, Any]]] = {}
    for rec in records:
        cat = rec.get("category", "기타")
        by_cat.setdefault(cat, []).append(rec)

    for cat in sorted(by_cat.keys()):
        ws = wb.create_sheet(cat[:31])  # 시트명 31자 제한
        _style_header(ws)
        for i, rec in enumerate(by_cat[cat], 2):
            row = _row_from_record(rec)
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=i, column=col_idx, value=val)
                cell.font = CELL_FONT
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            _grade_row_fill(ws, i, rec.get("main_grade", ""))
        _autosize(ws)

    # 메타 시트
    ws_meta = wb.create_sheet("Meta")
    ws_meta["A1"] = "SundayHug 전제품 네이버 SEO 키워드 전략"
    ws_meta["A1"].font = Font(name="Pretendard", size=14, bold=True, color="C8A07C")
    ws_meta["A3"] = "생성일"
    ws_meta["B3"] = date.today().strftime("%Y-%m-%d")
    ws_meta["A4"] = "총 제품 수"
    ws_meta["B4"] = len(records)
    ws_meta["A5"] = "S등급 키워드 수"
    ws_meta["B5"] = sum(1 for r in records if r.get("main_grade") == "S")
    ws_meta["A6"] = "A등급 키워드 수"
    ws_meta["B6"] = sum(1 for r in records if r.get("main_grade") == "A")
    ws_meta["A7"] = "주의 (경고 있는 제품)"
    ws_meta["B7"] = sum(1 for r in records if r.get("warnings"))
    ws_meta.column_dimensions["A"].width = 25
    ws_meta.column_dimensions["B"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"✅ xlsx 저장 완료: {output_path}")


if __name__ == "__main__":
    # 시범 데이터
    sample = [
        {
            "slug": "cooling-mesh-dual-pad",
            "category": "sleep-products",
            "title": "쿨링 메쉬 듀얼패드",
            "usp_short": "듀라론 냉감 × 메쉬 통기성",
            "material": "듀라론 / 3D 에어메쉬",
            "size": "S(100×70) / L(100×150)",
            "color": "아이보리",
            "product_name": "썬데이허그 꿀잠 아기 냉감패드 쿨매트 듀라론 메쉬 신생아 양면 여름",
            "main_keyword": "아기 냉감패드",
            "main_search": 20000,
            "main_grade": "S",
            "main_competition": "중간",
            "sub_keywords": ["아기 쿨매트", "신생아 냉감패드", "양면 쿨매트", "듀라론 패드"],
            "tags": [
                "아기냉감패드",
                "아기쿨매트",
                "신생아냉감패드",
                "듀라론패드",
                "양면쿨매트",
                "메쉬패드",
                "아기침대패드",
                "여름아기이불",
                "신생아쿨매트",
                "아기여름침구",
            ],
            "peak_month": 6,
            "current_vs_peak": "29.3%",
            "multi_categories": "",
            "warnings": [],
        }
    ]
    write_xlsx(sample, Path("/tmp/test-bulk-kw.xlsx"))
